import os
import openpyxl

def openFile(fN):
    with open(fN) as f:
        fL = [line.strip('\n') for line in f]
    return fL

book = openpyxl.Workbook()
for i in range(5, 12):
    fileInitial = ''.join(('datum', str(i), 'v2.txt'))
    ##fileAppend = ''.join(('datumAppend', str(i), '.txt'))
    fI = openFile(fileInitial)
    ##fA = openFile(fileAppend)
    sh = book.create_sheet(str(i))
    sh.cell(row=1,column=1).value = 'Distance'
    for j in range(0, 10):
        index = j * 102
        sh.cell(row=1, column=j+2).value = float(fI[index+1].split(' ')[1])
        for k in range(1, 101):
            #print j, index, k, fI[index + k]
            sh.cell(row=k+1,column=j+2).value = float(fI[index+k].split(' ')[0])
    ##sh.cell(row=1, column=11).value = float(fA[1].split(' ')[1])
    ##for l in range(1, 101):
        ##print j, index, k, fI[index + k]
        ##sh.cell(row=l+1,column=11).value = float(fA[l].split(' ')[0])
book.remove_sheet(book.get_sheet_by_name('Sheet'))
book.save(''.join(('BruteTSP2','.xlsx')))
